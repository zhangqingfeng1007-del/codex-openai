#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    import xlrd
except ImportError:  # pragma: no cover
    xlrd = None

try:
    import fitz
except ImportError:  # pragma: no cover
    fitz = None

try:
    import pdfplumber
except ImportError:  # pragma: no cover
    pdfplumber = None

from file_router import route_file


PAY_FREQ_ORDER = ["趸交", "年交", "半年交", "季交", "月交"]
PAY_PERIOD_ALIASES = {
    "趸交": "趸交",
    "一次交清": "趸交",
    "一次性付清": "趸交",
    "一次性交付": "趸交",
    "1年": "趸交",
}
RATE_SEARCH_ROOTS = [
    Path("/Users/zqf-openclaw/Desktop/开发材料/10款重疾"),
    Path("/Users/zqf-openclaw/Desktop/开发材料/招行数据"),
]


CHINESE_DIGITS = {
    "零": 0,
    "〇": 0,
    "一": 1,
    "二": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
}


def chinese_to_int(text: str) -> int | None:
    if not text:
        return None
    if text.isdigit():
        return int(text)
    if text == "十":
        return 10
    if "百" in text:
        left, right = text.split("百", 1)
        hundreds = CHINESE_DIGITS.get(left, 1 if left == "" else None)
        if hundreds is None:
            return None
        remainder = chinese_to_int(right) if right else 0
        if remainder is None:
            return None
        return hundreds * 100 + remainder
    if "十" in text:
        left, right = text.split("十", 1)
        tens = CHINESE_DIGITS.get(left, 1 if left == "" else None)
        if tens is None:
            return None
        ones = CHINESE_DIGITS.get(right, 0) if right else 0
        return tens * 10 + ones
    total = 0
    for ch in text:
        if ch not in CHINESE_DIGITS:
            return None
        total = total * 10 + CHINESE_DIGITS[ch]
    return total


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", "", text or "")


def iso_now() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def locate_rate_xlsx(item: dict) -> Path | None:
    clause_pdf_path = item.get("clause_pdf_path")
    product_id = item.get("product_id", "")
    if clause_pdf_path:
        product_dir = Path(clause_pdf_path).expanduser().resolve().parent
        if product_dir.exists():
            for pattern in ("*费率*.xlsx", "*费率*.xls"):
                candidates = sorted(product_dir.glob(pattern))
                if candidates:
                    return candidates[0]
    if product_id:
        for root in RATE_SEARCH_ROOTS:
            if not root.exists():
                continue
            for pattern in (f"*{product_id}*费率表.xlsx", f"*{product_id}*费率表.xls"):
                candidates = sorted(root.rglob(pattern))
                if candidates:
                    return candidates[0]
    return None


def locate_rate_pdf(item: dict) -> Path | None:
    clause_pdf_path = item.get("clause_pdf_path")
    product_id = item.get("product_id", "")
    if clause_pdf_path:
        product_dir = Path(clause_pdf_path).expanduser().resolve().parent
        if product_dir.exists():
            candidates = sorted(product_dir.glob("*费率*.pdf"))
            if candidates:
                return candidates[0]
    if product_id:
        for root in RATE_SEARCH_ROOTS:
            if not root.exists():
                continue
            candidates = sorted(root.rglob(f"*{product_id}*费率表.pdf"))
            if candidates:
                return candidates[0]
    return None


def locate_rate_source(item: dict) -> Path | None:
    path = locate_rate_xlsx(item)
    if path:
        return path
    return locate_rate_pdf(item)


def _normalize_pay_period_token(token: str) -> str | None:
    token = normalize_spaces(token)
    if not token:
        return None
    if token in PAY_PERIOD_ALIASES:
        return PAY_PERIOD_ALIASES[token]
    if token in {"一次交清", "一次性付清", "一次性交付"}:
        return "趸交"
    m = re.fullmatch(r"(\d+)年交", token)
    if m:
        years = int(m.group(1))
        return "趸交" if years == 1 else token
    m = re.fullmatch(r"(\d+)年", token)
    if m:
        years = int(m.group(1))
        return "趸交" if years == 1 else f"{years}年交"
    m = re.fullmatch(r"([一二三四五六七八九十百零〇]{1,4})年交", token)
    if m:
        years = chinese_to_int(m.group(1))
        if years is not None:
            return "趸交" if years == 1 else f"{years}年交"
    m = re.fullmatch(r"([一二三四五六七八九十百零〇]{1,4})年", token)
    if m:
        years = chinese_to_int(m.group(1))
        if years is not None:
            return "趸交" if years == 1 else f"{years}年交"
    m = re.fullmatch(r"交至(\d+)周岁", token)
    if m:
        return f"交至{m.group(1)}周岁"
    return None


def format_pay_periods(pay_periods: list[str]) -> list[str]:
    has_lump = False
    years: set[int] = set()
    age_targets: set[int] = set()
    for item in pay_periods:
        normalized = _normalize_pay_period_token(item)
        if not normalized:
            continue
        if normalized == "趸交":
            has_lump = True
            continue
        m = re.fullmatch(r"(\d+)年交", normalized)
        if m:
            years.add(int(m.group(1)))
            continue
        m = re.fullmatch(r"交至(\d+)周岁", normalized)
        if m:
            age_targets.add(int(m.group(1)))
    parts: list[str] = []
    if has_lump:
        parts.append("趸交")
    parts.extend(f"{y}年交" for y in sorted(years))
    parts.extend(f"交至{age}周岁" for age in sorted(age_targets))
    return parts


def _extract_pay_period_tokens(text: str) -> list[str]:
    compact = normalize_spaces(text)
    found: list[str] = []
    if any(k in compact for k in ["趸交", "一次交清", "一次性付清", "一次性交付"]):
        found.append("趸交")
    for raw in re.findall(r"(\d+)年交", compact):
        found.append(f"{raw}年交")
    for raw in re.findall(r"(\d+)年", compact):
        if raw == "1":
            found.append("趸交")
        else:
            found.append(f"{raw}年交")
    for raw in re.findall(r"([一二三四五六七八九十百零〇]{1,4})年交", compact):
        val = chinese_to_int(raw)
        if val is not None:
            found.append("趸交" if val == 1 else f"{val}年交")
    for raw in re.findall(r"([一二三四五六七八九十百零〇]{1,4})年", compact):
        val = chinese_to_int(raw)
        if val is not None:
            found.append("趸交" if val == 1 else f"{val}年交")
    for age in re.findall(r"交至(\d+)周岁", compact):
        found.append(f"交至{age}周岁")
    return format_pay_periods(found)


def _extract_pay_frequencies(text: str) -> list[str]:
    compact = normalize_spaces(text)
    compact = (
        compact.replace("一次交清", "趸交")
        .replace("一次性付清", "趸交")
        .replace("一次性交付", "趸交")
    )
    found = [freq for freq in PAY_FREQ_ORDER if freq in compact]
    mapped = [
        ("年缴", "年交"),
        ("每年", "年交"),
        ("半年缴", "半年交"),
        ("每半年", "半年交"),
        ("季缴", "季交"),
        ("每季", "季交"),
        ("月缴", "月交"),
        ("每月", "月交"),
    ]
    for token, freq in mapped:
        if token in compact and freq not in found:
            found.append(freq)
    return [freq for freq in PAY_FREQ_ORDER if freq in found]


def _extract_insurance_periods(text: str) -> list[str]:
    compact = normalize_spaces(text)
    periods: list[str] = []
    if "终身" in compact:
        periods.append("终身")
    for age in re.findall(r"至(\d+)周岁", compact):
        val = f"至{age}周岁"
        if val not in periods:
            periods.append(val)
    return periods


def _extract_pdf_pay_periods_by_lines(page_texts: list[str]) -> list[str]:
    candidates: list[str] = []
    for page_text in page_texts[:8]:
        lines = [line.strip() for line in page_text.splitlines() if line.strip()]
        if not any(line in {"缴费期", "交费期", "交费期间"} for line in lines):
            continue
        for item in lines:
            normalized = _normalize_pay_period_token(item)
            if normalized:
                candidates.append(normalized)
        # Some PDFs keep pay-period column headings as bare numbers (e.g. 10 / 20 / 30)
        # near the page tail instead of adjacent to "交费期间". Only use this heuristic
        # when no explicit period token was found from the page text itself.
        if not candidates:
            tail_numbers: list[str] = []
            for item in lines[-8:]:
                if re.fullmatch(r"\d{1,2}", item):
                    years = int(item)
                    if 2 <= years <= 60:
                        tail_numbers.append(f"{years}年交")
            if tail_numbers:
                candidates.extend(tail_numbers)
    return format_pay_periods(candidates)


def _fitz_page_texts(path: Path) -> tuple[list[str], str]:
    if fitz is None:
        return [], "fitz unavailable"
    try:
        doc = fitz.open(str(path))
    except Exception as exc:
        return [], f"fitz open failed: {exc}"
    try:
        texts = [page.get_text("text") or "" for page in doc]
        return texts, ""
    finally:
        doc.close()


def _pdfplumber_page_texts(path: Path) -> tuple[list[str], str]:
    if pdfplumber is None:
        return [], "pdfplumber unavailable"
    try:
        with pdfplumber.open(str(path)) as pdf:
            texts = [(page.extract_text() or "") for page in pdf.pages]
        return texts, ""
    except Exception as exc:
        return [], f"pdfplumber open failed: {exc}"


def _extract_pdf_texts(path: Path) -> tuple[list[str], str, list[str]]:
    warnings: list[str] = []
    texts, error = _fitz_page_texts(path)
    if texts:
        return texts, "fitz_text", warnings
    if error:
        warnings.append(error)
    texts, error = _pdfplumber_page_texts(path)
    if texts:
        return texts, "pdfplumber_text", warnings
    if error:
        warnings.append(error)
    return [], "unknown", warnings


def _build_rows_from_worksheet(sheet) -> list[dict]:
    rows_raw: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        values = [str(cell).strip() if cell is not None else "" for cell in row]
        if any(values):
            rows_raw.append(values)
    if not rows_raw:
        return []

    def _header_score(row: list[str]) -> tuple[int, int]:
        non_empty = [v for v in row if v]
        text_like = sum(1 for v in non_empty if re.search(r"[A-Za-z\u4e00-\u9fff]", v))
        return (text_like, len(non_empty))

    header_idx = max(range(min(len(rows_raw), 12)), key=lambda i: _header_score(rows_raw[i]))
    header = rows_raw[header_idx]
    body = rows_raw[header_idx + 1 : header_idx + 5]
    rows: list[dict] = []
    for body_row in body:
        row_obj: dict[str, str] = {}
        for idx, key in enumerate(header):
            if not key:
                continue
            row_obj[key] = body_row[idx] if idx < len(body_row) else ""
        if row_obj:
            rows.append(row_obj)
    return rows


def _extract_from_xls_xlrd(product_id: str, doc_category: str, input_file: Path) -> dict:
    """Extract rate table from legacy .xls files using xlrd."""
    if xlrd is None:
        raise RuntimeError("xlrd is required for .xls file extraction; install with: pip install xlrd")
    try:
        wb = xlrd.open_workbook(str(input_file))
    except Exception as exc:
        raise RuntimeError(f"xlrd failed to open {input_file}: {exc}") from exc

    tables: list[dict] = []
    pay_periods: list[str] = []
    pay_frequencies: list[str] = []
    insurance_periods: list[str] = []

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        rows_raw: list[list[str]] = []
        for row_idx in range(ws.nrows):
            row_vals = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    row_vals.append("")
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    # Avoid scientific notation for whole numbers
                    n = cell.value
                    row_vals.append(str(int(n)) if n == int(n) else str(n))
                else:
                    row_vals.append(str(cell.value).strip())
            if any(row_vals):
                rows_raw.append(row_vals)

        if not rows_raw:
            continue

        def _header_score(row: list[str]) -> tuple[int, int]:
            non_empty = [v for v in row if v]
            text_like = sum(1 for v in non_empty if re.search(r"[A-Za-z\u4e00-\u9fff]", v))
            return (text_like, len(non_empty))

        header_idx = max(range(min(len(rows_raw), 12)), key=lambda i: _header_score(rows_raw[i]))
        header = rows_raw[header_idx]
        body = rows_raw[header_idx + 1 : header_idx + 5]
        rows: list[dict] = []
        for body_row in body:
            row_obj: dict[str, str] = {}
            for idx, key in enumerate(header):
                if not key:
                    continue
                row_obj[key] = body_row[idx] if idx < len(body_row) else ""
            if row_obj:
                rows.append(row_obj)

        schema = list(rows[0].keys()) if rows else []
        tables.append({
            "table_id": f"{product_id}_{ws.name}_tbl_1",
            "page": 1,
            "schema": schema,
            "rows": rows,
        })

        # Extract pay periods from first 6 rows
        tokens: list[str] = []
        for row_idx in range(min(6, ws.nrows)):
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                val = ""
                if cell.ctype == xlrd.XL_CELL_NUMBER:
                    n = cell.value
                    val = str(int(n)) if n == int(n) else str(n)
                elif cell.ctype not in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    val = str(cell.value).strip()
                if val:
                    tokens.append(val)
                    normalized = _normalize_pay_period_token(val)
                    if normalized:
                        pay_periods.append(normalized)

        token_text = normalize_spaces(" ".join(tokens))
        if not pay_periods:
            pay_periods.extend(_extract_pay_period_tokens(token_text))

        # Extract pay frequencies from last 30 tokens
        all_parts: list[str] = []
        for row_idx in range(ws.nrows):
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                if cell.ctype not in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    val = str(cell.value).strip()
                    if val:
                        all_parts.append(val)
        tail_text = normalize_spaces("".join(all_parts[-30:]))
        pay_frequencies.extend(_extract_pay_frequencies(tail_text))
        insurance_periods.extend(_extract_insurance_periods(token_text))

    return {
        "_meta": {
            "product_id": product_id,
            "doc_category": doc_category,
            "source_file": str(input_file),
            "parse_method": "xlrd",
            "table_format": "sheet_grid",
            "parse_quality": "xls",
            "generated_at": iso_now(),
        },
        "tables": tables,
        "extracted_fields": {
            "pay_periods": format_pay_periods(pay_periods),
            "pay_frequencies": [freq for freq in PAY_FREQ_ORDER if freq in pay_frequencies],
            "insurance_periods": sorted(set(insurance_periods)),
        },
        "notes": [],
    }


def extract_from_xlsx(product_id: str, doc_category: str, input_file: Path) -> dict:
    if input_file.suffix.lower() == ".xls":
        return _extract_from_xls_xlrd(product_id, doc_category, input_file)
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for xlsx/xls extraction")
    workbook = load_workbook(input_file, read_only=True, data_only=True)
    try:
        tables: list[dict] = []
        pay_periods: list[str] = []
        pay_frequencies: list[str] = []
        insurance_periods: list[str] = []

        for sheet in workbook.worksheets:
            rows = _build_rows_from_worksheet(sheet)
            schema = list(rows[0].keys()) if rows else []
            tables.append(
                {
                    "table_id": f"{product_id}_{sheet.title}_tbl_1",
                    "page": 1,
                    "schema": schema,
                    "rows": rows,
                }
            )

            tokens: list[str] = []
            for row in sheet.iter_rows(min_row=1, max_row=6, values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    val = str(cell).strip()
                    if val:
                        tokens.append(val)

            token_text = normalize_spaces(" ".join(tokens))
            for token in tokens:
                normalized = _normalize_pay_period_token(str(token).strip())
                if normalized:
                    pay_periods.append(normalized)
            if not pay_periods:
                pay_periods.extend(_extract_pay_period_tokens(token_text))

            tail_parts: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    val = str(cell).strip()
                    if val:
                        tail_parts.append(val)
            tail_text = normalize_spaces("".join(tail_parts[-30:]))
            pay_frequencies.extend(_extract_pay_frequencies(tail_text))
            insurance_periods.extend(_extract_insurance_periods(token_text))

        return {
            "_meta": {
                "product_id": product_id,
                "doc_category": doc_category,
                "source_file": str(input_file),
                "parse_method": "openpyxl",
                "table_format": "sheet_grid",
                "parse_quality": "xlsx",
                "generated_at": iso_now(),
            },
            "tables": tables,
            "extracted_fields": {
                "pay_periods": format_pay_periods(pay_periods),
                "pay_frequencies": [freq for freq in PAY_FREQ_ORDER if freq in pay_frequencies],
                "insurance_periods": sorted(set(insurance_periods)),
            },
            "notes": [],
        }
    finally:
        workbook.close()


def extract_from_pdf(product_id: str, doc_category: str, input_file: Path) -> dict:
    texts, method, warnings = _extract_pdf_texts(input_file)
    if not texts:
        raise RuntimeError("; ".join(warnings) if warnings else "no pdf text extracted")
    full_text = "\n".join(texts)
    pay_periods = _extract_pdf_pay_periods_by_lines(texts)
    if not pay_periods:
        pay_periods = _extract_pay_period_tokens(full_text)
    pay_frequencies = _extract_pay_frequencies(full_text)
    insurance_periods = _extract_insurance_periods(full_text)

    notes = list(warnings)
    if "1年" in normalize_spaces(full_text):
        notes.append("缴费期列中的1年统一映射为趸交")

    first_page_lines = [line.strip() for line in texts[0].splitlines() if line.strip()]
    schema = []
    if first_page_lines:
        schema = first_page_lines[:3]

    tables = [
        {
            "table_id": f"{product_id}_rate_tbl_1",
            "page": 1,
            "schema": schema,
            "rows": [],
        }
    ]

    return {
        "_meta": {
            "product_id": product_id,
            "doc_category": doc_category,
            "source_file": str(input_file),
            "parse_method": method,
            "table_format": "rate_grid",
            "parse_quality": "text_pdf" if method in {"fitz_text", "pdfplumber_text"} else "unknown",
            "generated_at": iso_now(),
        },
        "tables": tables,
        "extracted_fields": {
            "pay_periods": pay_periods,
            "pay_frequencies": pay_frequencies,
            "insurance_periods": insurance_periods,
        },
        "notes": notes,
    }


def extract_rate_table(product_id: str, doc_category: str, input_file: Path, output_json: Path) -> dict:
    path = Path(input_file).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        payload = extract_from_xlsx(product_id, doc_category, path)
    elif suffix == ".pdf":
        payload = extract_from_pdf(product_id, doc_category, path)
    else:
        raise RuntimeError(f"unsupported rate file suffix: {suffix}")
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Extract raw rate/cash-value files to structured_table.json")
    parser.add_argument("--product-id", required=True)
    parser.add_argument("--doc-category", required=True, choices=["raw_rate", "cash_value"])
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--json", action="store_true", help="Print result as JSON")
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()
    input_path = Path(args.input).expanduser().resolve()

    route = route_file(input_path, args.doc_category)
    if route["extractor"] not in {"rate_table_extractor"}:
        parser.error(
            f"file routed to {route['extractor']} instead of rate_table_extractor: "
            f"{route['doc_category']} / {route['parse_quality']}"
        )

    payload = extract_rate_table(
        product_id=args.product_id,
        doc_category=args.doc_category,
        input_file=input_path,
        output_json=Path(args.output).expanduser().resolve(),
    )

    if args.json:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return

    print(
        f"wrote structured table for {args.product_id} to {Path(args.output).expanduser().resolve()}"
    )
    print(
        "extracted_fields:",
        json.dumps(payload.get("extracted_fields", {}), ensure_ascii=False),
    )


if __name__ == "__main__":
    main()
