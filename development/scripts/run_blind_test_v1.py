#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.append(str(SCRIPT_DIR))

from parse_md_blocks import parse_markdown  # noqa: E402
from extract_tier_a_rules import extract_candidates  # noqa: E402


CORE_FIELDS = [
    "投保年龄",
    "保险期间",
    "交费期间",
    "交费频率",
    "等待期",
    "等待期（简化）",
    "宽限期",
    "犹豫期",
    "重疾赔付次数",
    "重疾分组",
    "重疾数量",
]

RATE_FIRST_FIELDS = {"投保年龄", "交费期间", "交费频率"}


def normalize_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def rate_sort_key(value: str) -> tuple[int, int | str]:
    if value == "趸交":
        return (0, 0)
    if value.endswith("年"):
        try:
            return (1, int(value[:-1]))
        except ValueError:
            return (1, value)
    if value.endswith("年交"):
        try:
            return (2, int(value[:-2]))
        except ValueError:
            return (2, value)
    order = {"年交": 3, "半年交": 4, "季交": 5, "月交": 6}
    return (order.get(value, 99), value)


def convert_clause_to_md(clause_path: Path, output_dir: Path, engine: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    expected = output_dir / f"{clause_path.stem}.md"
    if expected.exists():
        return expected

    venv_python = Path("/Users/zqf-openclaw/codex-openai/development/.venv_doc_to_md/bin/python")
    with tempfile.TemporaryDirectory(prefix="blind_test_clause_") as temp_dir:
        temp_input_dir = Path(temp_dir)
        temp_input_file = temp_input_dir / clause_path.name
        shutil.copy2(clause_path, temp_input_file)
        cmd = [
            str(venv_python),
            "-m",
            "doc_to_md.cli",
            "convert",
            "--input-path",
            str(temp_input_dir),
            "--output-path",
            str(output_dir),
            "--engine",
            engine,
        ]
        subprocess.run(cmd, check=True, cwd="/Users/zqf-openclaw/codex-openai/development/vendor/doc_to_md")
    if not expected.exists():
        raise FileNotFoundError(f"Markdown not generated: {expected}")
    return expected


def load_processed_rate_candidates(rate_files: list[str]) -> dict[str, dict[str, Any]]:
    ages: list[int] = []
    pay_periods: set[str] = set()
    pay_freqs: set[str] = set()
    insurance_periods: set[str] = set()
    used_files: list[str] = []

    for file_path in rate_files:
        if not file_path or "中转表" in file_path:
            continue
        path = Path(file_path)
        if not path.exists():
            continue
        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [normalize_value(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        idx = {name: i for i, name in enumerate(headers)}

        age_idx = idx.get("投保年龄")
        pay_period_idx = idx.get("缴费期间") if "缴费期间" in idx else idx.get("交费期间")
        insurance_idx = idx.get("保险期间")
        pay_freq_idx = idx.get("交费方式")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if age_idx is not None and age_idx < len(row):
                raw_age = normalize_value(row[age_idx])
                if raw_age.isdigit():
                    ages.append(int(raw_age))
            if pay_period_idx is not None and pay_period_idx < len(row):
                raw = normalize_value(row[pay_period_idx])
                if raw:
                    pay_periods.add(raw)
            if insurance_idx is not None and insurance_idx < len(row):
                raw = normalize_value(row[insurance_idx])
                if raw:
                    insurance_periods.add(raw)
            if pay_freq_idx is not None and pay_freq_idx < len(row):
                raw = normalize_value(row[pay_freq_idx])
                if raw:
                    pay_freqs.add(raw)
        used_files.append(str(path))

    results: dict[str, dict[str, Any]] = {}
    primary_file = used_files[0] if used_files else None

    if ages and primary_file:
        results["投保年龄"] = {
            "coverage_name": "投保年龄",
            "value": f"{min(ages)}-{max(ages)}周岁",
            "confidence": 0.98,
            "status": "candidate_ready",
            "source_type": "processed_rate",
            "source_file": primary_file,
            "evidence": {
                "summary": f"投保年龄列最小={min(ages)}，最大={max(ages)}",
            },
            "extract_method": "processed_rate:distinct_age_range",
        }

    if pay_periods and primary_file:
        ordered = sorted(pay_periods, key=rate_sort_key)
        results["交费期间"] = {
            "coverage_name": "交费期间",
            "value": "，".join(ordered),
            "confidence": 0.98,
            "status": "candidate_ready",
            "source_type": "processed_rate",
            "source_file": primary_file,
            "evidence": {
                "summary": f"缴费期间去重值={ordered}",
            },
            "extract_method": "processed_rate:distinct_pay_periods",
        }

    if pay_freqs and primary_file:
        ordered = sorted(pay_freqs, key=rate_sort_key)
        results["交费频率"] = {
            "coverage_name": "交费频率",
            "value": "，".join(ordered),
            "confidence": 0.98,
            "status": "candidate_ready",
            "source_type": "processed_rate",
            "source_file": primary_file,
            "evidence": {
                "summary": f"交费方式去重值={ordered}",
            },
            "extract_method": "processed_rate:distinct_pay_frequency",
        }

    if insurance_periods and primary_file:
        ordered = sorted(insurance_periods, key=rate_sort_key)
        results["保险期间"] = {
            "coverage_name": "保险期间",
            "value": "，".join(ordered),
            "confidence": 0.96,
            "status": "candidate_ready",
            "source_type": "processed_rate",
            "source_file": primary_file,
            "evidence": {
                "summary": f"保险期间去重值={ordered}",
            },
            "extract_method": "processed_rate:distinct_insurance_period",
        }

    return results


def decorate_clause_candidates(candidates: dict[str, dict[str, Any]], clause_path: str, md_path: str) -> dict[str, dict[str, Any]]:
    decorated: dict[str, dict[str, Any]] = {}
    for field, candidate in candidates.items():
        entry = dict(candidate)
        entry.setdefault("status", "candidate_ready")
        entry["source_type"] = "clause"
        entry["source_file"] = clause_path
        entry["source_md_file"] = md_path
        entry["evidence"] = {
            "block_id": candidate.get("block_id"),
            "page": candidate.get("page"),
            "text": candidate.get("evidence_text"),
        }
        entry["extract_method"] = candidate.get("note")
        decorated[field] = entry
    return decorated


def merge_candidates(clause_candidates: dict[str, dict[str, Any]], rate_candidates: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    merged = dict(clause_candidates)
    for field, candidate in rate_candidates.items():
        current = merged.get(field)
        if field in RATE_FIRST_FIELDS:
            merged[field] = candidate
            continue
        if current is None:
            merged[field] = candidate
    return merged


def run_product(product: dict[str, Any], md_dir: Path) -> dict[str, Any]:
    clause_path = Path(product["source_files"]["clause"][0])
    md_path = convert_clause_to_md(clause_path, md_dir, engine="markitdown")
    blocks = parse_markdown(product["product_id"], md_path.read_text(encoding="utf-8"))
    clause_candidates = decorate_clause_candidates(
        extract_candidates(blocks),
        clause_path=str(clause_path),
        md_path=str(md_path),
    )
    rate_candidates = load_processed_rate_candidates(product["source_files"].get("processed_rate", []))
    merged = merge_candidates(clause_candidates, rate_candidates)

    items = [merged[field] for field in CORE_FIELDS if field in merged]
    missing = [field for field in CORE_FIELDS if field not in merged]
    return {
        "product_id": product["product_id"],
        "product_name": product["product_name"],
        "difficulty": product["difficulty"],
        "items": items,
        "missing_core_fields": missing,
        "used_sources": {
            "clause": product["source_files"].get("clause", []),
            "processed_rate": product["source_files"].get("processed_rate", []),
            "product_brochure": product["source_files"].get("product_brochure", []),
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Run blind test V1 on 10 products.")
    parser.add_argument("--manifest", required=True)
    parser.add_argument("--output-dir", required=True)
    args = parser.parse_args()

    manifest = json.loads(Path(args.manifest).read_text(encoding="utf-8"))
    output_dir = Path(args.output_dir)
    md_dir = output_dir / "md_cache"
    product_dir = output_dir / "products"
    product_dir.mkdir(parents=True, exist_ok=True)

    results = []
    for product in manifest:
        if product.get("status") != "ready":
            continue
        result = run_product(product, md_dir)
        results.append(result)
        (product_dir / f"{product['product_id']}.json").write_text(
            json.dumps(result, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    summary = {
        "run_date": "2026-03-25",
        "product_count": len(results),
        "products": results,
    }
    (output_dir / "blind_test_results_v1.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps({"product_count": len(results), "output": str(output_dir / "blind_test_results_v1.json")}, ensure_ascii=False))


if __name__ == "__main__":
    main()
