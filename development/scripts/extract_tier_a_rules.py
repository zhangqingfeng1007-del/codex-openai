#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from block_tagger import tag_blocks

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency in local env
    load_workbook = None

try:
    import fitz
except ImportError:  # pragma: no cover - optional dependency in local env
    fitz = None


TARGET_FIELDS = [
    "投保年龄",
    "保险期间",
    "交费期间",
    "交费频率",
    "等待期",
    "等待期（简化）",
    "宽限期",
    "犹豫期",
    "重疾赔付次数",
    "重疾分组",
    "重疾数量",
]


CHINESE_DIGITS = {
    "零": 0,
    "〇": 0,
    "一": 1,
    "二": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
}

PAY_FREQ_ORDER = ["趸交", "年交", "半年交", "季交", "月交"]
WAITING_PRIORITY_TITLES = ("等待期", "观察期", "投保须知")
PAY_PERIOD_TITLES = [
    "交费期间", "交费期限", "交费年期",
    "缴费期间", "缴费期限", "缴费年期",
    "保险费缴纳期间", "缴费方式",
]
RATE_SEARCH_ROOTS = [
    Path("/Users/zqf-openclaw/Desktop/开发材料/10款重疾"),
    Path("/Users/zqf-openclaw/Desktop/开发材料/招行数据"),
]
RATE_TABLES_DIR = Path(__file__).resolve().parents[1] / "data" / "tables"


def load_blocks_compatible(path: Path) -> tuple[dict | None, list[dict]]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(raw, dict) and "blocks" in raw:
        return raw.get("_meta"), raw["blocks"]
    if isinstance(raw, list):
        return None, raw
    raise ValueError(f"unsupported blocks payload: {path}")


def load_manifest_compatible(path: Path) -> list[dict]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, list):
        raise ValueError(f"unsupported manifest payload: {path}")

    normalized: list[dict] = []
    for item in raw:
        if not isinstance(item, dict):
            raise ValueError(f"unsupported manifest row: {item!r}")

        if "files" in item:
            normalized.append(item)
            continue

        clause_pdf_path = item.get("clause_pdf_path")
        files = []
        if clause_pdf_path:
            files.append(
                {
                    "doc_category": "clause",
                    "source_file": clause_pdf_path,
                    "extractor": "pdf_text_extractor",
                    "is_raw": True,
                }
            )

        normalized_item = dict(item)
        normalized_item["files"] = files
        normalized.append(normalized_item)

    return normalized


def load_structured_table_compatible(path: Path) -> tuple[dict | None, list[dict], dict]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError(f"unsupported structured table payload: {path}")
    return raw.get("_meta"), raw.get("tables", []), raw.get("extracted_fields", {})


def locate_structured_table_json(item: dict, doc_category: str = "raw_rate") -> Path | None:
    product_id = item.get("product_id", "")
    candidates: list[Path] = []

    files = item.get("files") or []
    for file_entry in files:
        if file_entry.get("doc_category") != doc_category:
            continue
        source_file = file_entry.get("source_file")
        if not source_file:
            continue
        stem = Path(source_file).stem
        candidates.append(RATE_TABLES_DIR / f"{stem}_structured_table.json")

    if product_id:
        suffix = "rate" if doc_category == "raw_rate" else doc_category
        candidates.extend(
            [
                RATE_TABLES_DIR / f"{product_id}_{suffix}_structured_table.json",
                RATE_TABLES_DIR / f"{product_id}_{doc_category}_structured_table.json",
                RATE_TABLES_DIR / f"{product_id}_structured_table.json",
            ]
        )

    seen: set[Path] = set()
    for candidate in candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        if candidate.exists():
            return candidate
    return None


def _format_pay_period_candidate(values: list[str]) -> str | None:
    if not values:
        return None
    has_lump = any(v == "趸交" for v in values)
    year_values: set[int] = set()
    age_periods: list[str] = []
    for value in values:
        m = re.fullmatch(r"(\d+)年交", value)
        if m:
            y = int(m.group(1))
            if y <= 40:
                year_values.add(y)
            continue
        if value.startswith("交至") and value not in age_periods:
            age_periods.append(value)
    parts: list[str] = []
    if has_lump:
        parts.append("趸交")
    if year_values:
        parts.append("/".join(str(y) for y in sorted(year_values)) + "年交")
    parts.extend(age_periods)
    return "，".join(parts) if parts else None


def build_candidates_from_structured_table(path: Path) -> dict[str, dict]:
    meta, _, extracted_fields = load_structured_table_compatible(path)
    source_file = meta.get("source_file") if meta else str(path)
    parse_method = meta.get("parse_method") if meta else "structured_table"
    parse_quality = meta.get("parse_quality") if meta else "unknown"
    note_suffix = f"[{Path(source_file).name}]"

    candidates: dict[str, dict] = {}

    pay_periods = extracted_fields.get("pay_periods") or []
    if isinstance(pay_periods, list):
        value = _format_pay_period_candidate([str(v) for v in pay_periods if v])
        if value:
            candidates["交费期间"] = {
                "coverage_name": "交费期间",
                "value": value,
                "confidence": 0.84,
                "note": f"rule: structured_table_pay_period{note_suffix}",
                "block_id": None,
                "page": None,
                "evidence_text": source_file,
                "source_type": "structured_table",
                "parse_method": parse_method,
                "parse_quality": parse_quality,
            }

    pay_frequencies = extracted_fields.get("pay_frequencies") or []
    if isinstance(pay_frequencies, list):
        ordered = [freq for freq in PAY_FREQ_ORDER if freq in pay_frequencies]
        if ordered:
            candidates["交费频率"] = {
                "coverage_name": "交费频率",
                "value": "，".join(ordered),
                "confidence": 0.84,
                "note": f"rule: structured_table_pay_frequency{note_suffix}",
                "block_id": None,
                "page": None,
                "evidence_text": source_file,
                "source_type": "structured_table",
                "parse_method": parse_method,
                "parse_quality": parse_quality,
            }

    insurance_periods = extracted_fields.get("insurance_periods") or []
    if isinstance(insurance_periods, list):
        unique_periods: list[str] = []
        for value in insurance_periods:
            text = str(value).strip()
            if text and text not in unique_periods:
                unique_periods.append(text)
        if unique_periods:
            # Normalize order: "至X周岁" before "终身" (matches standard value format)
            def _period_sort_key(p: str) -> int:
                if p.startswith("至") and "周岁" in p:
                    return 0
                if p == "终身":
                    return 1
                return 2
            unique_periods.sort(key=_period_sort_key)
            candidates["保险期间"] = {
                "coverage_name": "保险期间",
                "value": "，".join(unique_periods),
                "confidence": 0.83,
                "note": f"rule: structured_table_insurance_period{note_suffix}",
                "block_id": None,
                "page": None,
                "evidence_text": source_file,
                "source_type": "structured_table",
                "parse_method": parse_method,
                "parse_quality": parse_quality,
            }

    return candidates


def chinese_to_int(text: str) -> int | None:
    if not text:
        return None
    if text.isdigit():
        return int(text)
    if text == "十":
        return 10
    if "百" in text:
        left, right = text.split("百", 1)
        hundreds = CHINESE_DIGITS.get(left, 1 if left == "" else None)
        if hundreds is None:
            return None
        remainder = chinese_to_int(right) if right else 0
        if remainder is None:
            return None
        return hundreds * 100 + remainder
    if "十" in text:
        left, right = text.split("十", 1)
        tens = CHINESE_DIGITS.get(left, 1 if left == "" else None)
        if tens is None:
            return None
        ones = CHINESE_DIGITS.get(right, 0) if right else 0
        return tens * 10 + ones
    total = 0
    for ch in text:
        if ch not in CHINESE_DIGITS:
            return None
        total = total * 10 + CHINESE_DIGITS[ch]
    return total


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", "", text)


def narrow_evidence(block_text: str, keyword: str, window: int = 150) -> str:
    """Extract minimal snippet around first occurrence of keyword in block_text."""
    idx = block_text.find(keyword)
    if idx == -1:
        return block_text[:200]
    # Start: find nearest sentence boundary (。or newline) before keyword
    start = max(0, idx - 60)
    for sep in ["。", "\n"]:
        pos = block_text.rfind(sep, 0, idx)
        if pos != -1 and pos + 1 > start:
            start = pos + 1
    # End: find nearest sentence boundary after keyword
    end = min(len(block_text), idx + window)
    for sep in ["。", "\n"]:
        pos = block_text.find(sep, idx)
        if pos != -1 and pos + 1 < end:
            end = pos + 1
    return block_text[start:end].strip()


def add_candidate(results: dict, field: str, value: str | None, block: dict | None, confidence: float, note: str) -> None:
    if not value:
        return
    if confidence <= 0:
        return
    entry = {
        "coverage_name": field,
        "value": value,
        "confidence": round(confidence, 2),
        "note": note,
        "block_id": block["block_id"] if block else None,
        "page": block["page"] if block else None,
        "evidence_text": block["text"] if block else None,
    }
    current = results.get(field)
    if current is None or entry["confidence"] > current["confidence"]:
        results[field] = entry


def set_candidate(results: dict, field: str, value: str, block: dict | None, confidence: float, note: str, status: str | None = None) -> None:
    entry = {
        "coverage_name": field,
        "value": value,
        "confidence": round(confidence, 2),
        "note": note,
        "block_id": block.get("block_id") if block else None,
        "page": block.get("page") if block else None,
        "evidence_text": block.get("text") or block.get("evidence_text") if block else None,
    }
    if status:
        entry["status"] = status
    results[field] = entry


def extract_age(text: str) -> str | None:
    compact = normalize_spaces(text)
    m = re.search(r"(?:投保年龄.*?为|出生满)(\d+)(?:天|日).*?(\d+)周岁", compact)
    if m:
        return f"0（{m.group(1)}天）-{m.group(2)}周岁"
    m = re.search(r"(?:投保年龄.*?为|接受的投保年龄为)?(\d+)周岁至(\d+)周岁", compact)
    if m:
        return f"{m.group(1)}-{m.group(2)}周岁"
    m = re.search(r"出生([一二三四五六七八九十百零〇]+)日以上.*?([一二三四五六七八九十百零〇\d]+)周岁以下", compact)
    if m:
        low = chinese_to_int(m.group(1))
        high = chinese_to_int(m.group(2))
        if low is not None and high is not None:
            return f"0（{low}天）-{high}周岁"
    m = re.search(r"([一二三四五六七八九十百零〇\d]+)周岁以下", compact)
    if "投保年龄" in compact and m:
        high = chinese_to_int(m.group(1))
        if high is not None:
            return f"-{high}周岁"
    return None


def extract_duration_days(text: str, keyword: str, max_days: int | None = None) -> str | None:
    """Extract a day-count near *keyword* in *text*.

    Searches in a context window (40 chars before … 100 chars after the first
    occurrence of *keyword*) so that unrelated numbers elsewhere in the same
    block do not pollute the result.  Recognises "N个自然日" as well as "N天/日".
    If *max_days* is given, values exceeding that limit are rejected.
    """
    compact = normalize_spaces(text)
    if keyword not in compact:
        return None
    pos = compact.find(keyword)
    # Search after the keyword first (preferred), fall back to a broader window
    # that includes up to 40 chars before the keyword.
    after = compact[pos : pos + 120]
    before_and_after = compact[max(0, pos - 40) : pos + 120]

    def _find_days(window: str) -> int | None:
        m = re.search(r"(\d+)个自然日", window)
        if not m:
            m = re.search(r"(\d+)[天日]", window)
        if m:
            return int(m.group(1))
        m_cn = re.search(r"([一二三四五六七八九十百零〇]+)[天日]", window)
        if m_cn:
            return chinese_to_int(m_cn.group(1))
        return None

    val = _find_days(after)
    if val is None:
        val = _find_days(before_and_after)
    if val is None:
        return None
    if max_days is not None and val > max_days:
        return None
    return f"{val}天"


def extract_waiting(text: str) -> tuple[str | None, str | None]:
    compact = normalize_spaces(text)
    if "等待期" not in compact and "观察期" not in compact:
        return None, None
    m = re.search(r"(\d+)[天日]内", compact)
    if not m:
        m = re.search(r"(\d+)[天日]", compact)
    if not m:
        m_cn = re.search(r"([一二三四五六七八九十百零〇]+)[天日]", compact)
        if m_cn:
            val = chinese_to_int(m_cn.group(1))
            if val is not None:
                days = str(val)
                simplified = f"{days}天"
                if "意外" in compact:
                    return f"非意外{days}天，意外0天", simplified
                return simplified, simplified
        return None, None
    days = m.group(1)
    simplified = f"{days}天"
    if "意外" in compact:
        return f"非意外{days}天，意外0天", simplified
    return simplified, simplified


def should_skip_waiting_block(block: dict) -> bool:
    """Use _tags from block_tagger when available, else fall back to inline check."""
    tags = block.get("_tags")
    if tags:
        return not tags["is_matchable"]
    # Legacy fallback (should not be reached after tag_blocks is wired)
    title_path = "".join(block.get("title_path", []))
    text = block.get("text", "")
    skip_titles = ("举例", "示例", "案例", "理赔案例", "保险金给付案例")
    skip_texts = ("小王", "小明", "案例", "举例说明", "举例")
    return any(hint in title_path for hint in skip_titles) or any(hint in text for hint in skip_texts)


def waiting_block_confidence(block: dict) -> float:
    if should_skip_waiting_block(block):
        return 0.0
    text = block["text"]
    # Content-level skip: calculation context, not actual waiting period definition
    if "确诊日起满" in text or "保单周年日" in text:
        return 0.0
    title_path = "".join(block.get("title_path", []))
    if any(keyword in title_path for keyword in WAITING_PRIORITY_TITLES):
        return 0.97
    # 正文直接含等待期/观察期定义（如"等待期为X天""X天等待期"）→ 高置信
    compact_text = normalize_spaces(text)
    if re.search(r"(?:等待期|观察期)[^。]{0,50}\d+[天日]", compact_text) or \
       re.search(r"\d+[天日][^。]{0,50}(?:等待期|观察期)", compact_text):
        return 0.98
    return 0.93


def extract_insurance_period(text: str) -> str | None:
    compact = normalize_spaces(text)
    if "保险期间" not in compact and "保障期间" not in compact:
        return None
    m = re.search(r"终身[，,]或.*?至被保险人([一二三四五六七八九十百]+)周岁", compact)
    if m:
        age = chinese_to_int(m.group(1))
        if age is not None:
            return f"至{age}周岁，终身"
    m = re.search(r"至被保险人([一二三四五六七八九十百]+)周岁.*?终身", compact)
    if m:
        age = chinese_to_int(m.group(1))
        if age is not None:
            return f"至{age}周岁，终身"
    m = re.search(r"至(\d+)周岁[，,]终身", compact)
    if m:
        return f"至{m.group(1)}周岁，终身"
    m = re.search(r"终身[或和及/]至(\d+)周岁", compact)
    if m:
        return f"至{m.group(1)}周岁，终身"
    m = re.search(r"至(\d+)周岁", compact)
    age_part = f"至{m.group(1)}周岁" if m else None
    m = re.search(r"保(?:至|到)(\d+)岁", compact)
    if not age_part and m:
        age_part = f"至{m.group(1)}周岁"
    has_lifetime = "终身" in compact
    if age_part and has_lifetime:
        return f"{age_part}，终身"
    if has_lifetime:
        return "终身"
    if age_part:
        return age_part
    # 支持"至被保险人 25 周岁"（阿拉伯数字）
    m = re.search(r"至被保险人\s*(\d+)\s*周岁", compact)
    if m:
        age_part = f"至{m.group(1)}周岁"
        if "终身" in compact:
            return f"{age_part}，终身"
        return age_part
    return None


def _format_pay_periods_from_text(compact: str) -> str | None:
    freqs: list[str] = []
    if any(k in compact for k in ["趸交", "一次性交付", "一次交清", "一次性付清"]):
        freqs.append("趸交")
    compact_clean = re.sub(r"(\d{1,2})年交\1(?=\d{1,2}年交)", r"\1年交", compact)
    compact_clean = re.sub(r"(\d{1,2})年交\1(?=[^\d]|$)", r"\1年交", compact_clean)
    years: set[int] = {int(x) for x in re.findall(r"(\d+)年交", compact_clean)}
    for raw in re.findall(r"([一二三四五六七八九十百零〇]{1,4})年交", compact):
        val = chinese_to_int(raw)
        if val is not None:
            years.add(val)
    if not years:
        m = re.search(
            r"交费期间(?:分为|包括)(.+?)(?:两种|三种|四种|五种|六种|七种|八种|九种|十种|，|。|交费方式)",
            compact,
        )
        if m:
            segment = m.group(1)
            years |= {int(x) for x in re.findall(r"(\d+)年", segment)}
            for raw in re.findall(r"([一二三四五六七八九十百零〇]{1,4})年", segment):
                val = chinese_to_int(raw)
                if val is not None:
                    years.add(val)
    if years:
        freqs.append("/".join(str(y) for y in sorted(years)) + "年交")
    age_match = re.search(r"交至(\d+)周岁", compact)
    if age_match:
        freqs.append(f"交至{age_match.group(1)}周岁")
    if not freqs:
        return None
    ordered: list[str] = []
    for item in freqs:
        if item not in ordered:
            ordered.append(item)
    return "，".join(ordered)


def extract_pay_period(text: str) -> str | None:
    compact = normalize_spaces(text)
    if not any(k in compact for k in PAY_PERIOD_TITLES + ["交费方式和交费期间", "保险费的支付"]):
        return None
    return _format_pay_periods_from_text(compact)


def extract_pay_frequency(text: str) -> str | None:
    compact = normalize_spaces(text)
    definition_like_patterns = [
        "根据本合同交费方式确定",
        "根据交费方式确定",
        "对应日",
        "对应的日期",
        "每月第一个营业日",
        "根据市场情况",
        "约定利率",
        "每月每季每半年或每年对应的日",
        "保单周年日",
        "保险费约定支付日",
        "生效对应日",
        "每年至少一次",
        "合同生效后每年的对应日",
        "合同生效后每月的对应日",
        "合同生效日在合同生效后",
    ]
    normalized_compact = compact.replace("、", "").replace("，", "").replace("（", "").replace("）", "")
    if any(pat.replace("、", "").replace("，", "").replace("（", "").replace("）", "") in normalized_compact for pat in definition_like_patterns):
        return None
    if not any(k in compact for k in ["交费频率", "交费方式", "保险费的支付"]):
        if not any(k in compact for k in ["每月", "每季", "每半年", "每年", "趸交", "一次交清", "一次性付清"]):
            return None
    found = [freq for freq in PAY_FREQ_ORDER if freq in compact]
    mapped = [
        ("每月", "月交"),
        ("每季", "季交"),
        ("每半年", "半年交"),
        ("每年", "年交"),
        ("一次交清", "趸交"),
        ("一次性付清", "趸交"),
        ("一次性交付", "趸交"),
        ("趸缴", "趸交"),
        ("年缴", "年交"),
        ("半年缴", "半年交"),
        ("季缴", "季交"),
        ("月缴", "月交"),
    ]
    for token, freq in mapped:
        if token in compact and freq not in found:
            found.append(freq)
    found = [freq for freq in PAY_FREQ_ORDER if freq in found]
    return "，".join(found) if found else None


def locate_rate_xlsx(item: dict) -> Path | None:
    clause_pdf_path = item.get("clause_pdf_path")
    product_id = item.get("product_id", "")
    if clause_pdf_path:
        product_dir = Path(clause_pdf_path).expanduser().resolve().parent
        if product_dir.exists():
            for pattern in ("*费率*.xlsx", "*费率*.xls"):
                candidates = sorted(product_dir.glob(pattern))
                if candidates:
                    return candidates[0]
    if product_id:
        for root in RATE_SEARCH_ROOTS:
            if not root.exists():
                continue
            for pattern in (f"*{product_id}*费率表.xlsx", f"*{product_id}*费率表.xls"):
                candidates = sorted(root.rglob(pattern))
                if candidates:
                    return candidates[0]
    return None


def locate_rate_pdf(item: dict) -> Path | None:
    clause_pdf_path = item.get("clause_pdf_path")
    product_id = item.get("product_id", "")
    if clause_pdf_path:
        product_dir = Path(clause_pdf_path).expanduser().resolve().parent
        if product_dir.exists():
            candidates = sorted(product_dir.glob("*费率*.pdf"))
            if candidates:
                return candidates[0]
    if product_id:
        for root in RATE_SEARCH_ROOTS:
            if not root.exists():
                continue
            candidates = sorted(root.rglob(f"*{product_id}*费率表.pdf"))
            if candidates:
                return candidates[0]
    return None


def extract_pay_info_from_rate_pdf(item: dict) -> dict[str, str]:
    if fitz is None:
        return {}
    rate_path = locate_rate_pdf(item)
    if rate_path is None or not rate_path.exists():
        return {}
    try:
        doc = fitz.open(rate_path)
    except Exception:
        return {}

    try:
        text = "\n".join(page.get_text("text") for page in doc)
    finally:
        doc.close()

    compact = normalize_spaces(text)
    out: dict[str, str] = {}

    # Build pay-period set: limit to 1–2 digit year counts to avoid noise from
    # decimal values (e.g. "0.97362..." → "97362年交" in the old regex).
    years: set[int] = set()
    has_lump = any(k in compact for k in ["趸交", "一次交清", "一次性付清", "一次性交付", "趸缴"])
    # "NX年交" (explicit) takes priority
    for x in re.findall(r"(?<!\d)(\d{1,2})年交(?!\d)", compact):
        val = int(x)
        if val == 1:
            has_lump = True
        elif 2 <= val <= 99:
            years.add(val)
    # Plain "NX年" (implicit, only if no "交" immediately follows)
    for x in re.findall(r"(?<!\d)(\d{1,2})年(?!交)(?!\d)", compact):
        val = int(x)
        if val == 1:
            has_lump = True
        elif 2 <= val <= 99:
            years.add(val)
    if has_lump or years:
        parts: list[str] = []
        if has_lump:
            parts.append("趸交")
        if years:
            parts.append("/".join(str(y) for y in sorted(years)) + "年交")
        out["交费期间"] = "，".join(parts)

    # Detect pay frequencies: support both "缴" and "交" terminology
    freqs: list[str] = []
    freq_tokens = [
        ("趸交", "趸交"), ("趸缴", "趸交"),
        ("年缴", "年交"), ("年交", "年交"),
        ("半年缴", "半年交"), ("半年交", "半年交"),
        ("季缴", "季交"), ("季交", "季交"),
        ("月缴", "月交"), ("月交", "月交"),
    ]
    for token, freq in freq_tokens:
        if token in compact and freq not in freqs:
            freqs.append(freq)
    if has_lump and "趸交" not in freqs:
        freqs.append("趸交")
    if freqs:
        ordered = [freq for freq in PAY_FREQ_ORDER if freq in freqs]
        out["交费频率"] = "，".join(ordered)
    return out


def extract_age_from_rate_source(item: dict) -> str | None:
    """Try to extract 投保年龄 from rate XLSX or PDF header rows.

    Rate tables often have cells like "承保年龄：0-60周岁" or
    "投保年龄 出生满28天-65周岁" in the first 10 rows.
    """
    # Try XLSX first
    if load_workbook is not None:
        rate_path = locate_rate_xlsx(item)
        if rate_path and rate_path.exists():
            try:
                wb = load_workbook(rate_path, read_only=True, data_only=True)
                try:
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                            for cell in row:
                                if cell is None:
                                    continue
                                val = str(cell).strip()
                                age = extract_age(val)
                                if age:
                                    return age
                                # "承保年龄" / "投保年龄" label followed by range in same cell
                                compact_cell = normalize_spaces(val)
                                if any(k in compact_cell for k in ["承保年龄", "投保年龄"]):
                                    age = extract_age(compact_cell)
                                    if age:
                                        return age
                                    # Try "N-M周岁" or "N至M周岁" pattern directly
                                    m = re.search(r"(\d+)[至-](\d+)周岁", compact_cell)
                                    if m:
                                        return f"{m.group(1)}-{m.group(2)}周岁"
                finally:
                    wb.close()
            except Exception:
                pass

    # Try PDF
    if fitz is not None:
        rate_path = locate_rate_pdf(item)
        if rate_path and rate_path.exists():
            try:
                doc = fitz.open(str(rate_path))
                try:
                    for page in doc[:3]:
                        text = page.get_text("text") or ""
                        for line in text.splitlines():
                            line = line.strip()
                            if not line:
                                continue
                            compact_line = normalize_spaces(line)
                            if any(k in compact_line for k in ["承保年龄", "投保年龄"]):
                                age = extract_age(compact_line)
                                if age:
                                    return age
                                m = re.search(r"(\d+)[至-](\d+)周岁", compact_line)
                                if m:
                                    return f"{m.group(1)}-{m.group(2)}周岁"
                finally:
                    doc.close()
            except Exception:
                pass
    return None


def extract_pay_frequency_from_rate_xlsx(item: dict) -> tuple[str, str] | None:
    if load_workbook is None:
        return None
    rate_path = locate_rate_xlsx(item)
    if rate_path is None or not rate_path.exists():
        return None
    try:
        workbook = load_workbook(rate_path, read_only=True, data_only=True)
    except Exception:
        return None

    try:
        for sheet in workbook.worksheets:
            # max_row may be 0/None in read_only mode for some xlsx — collect all rows then take last 5
            all_parts: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell).strip()
                    if text:
                        all_parts.append(text)
            # Use the last 30 tokens (footer area)
            tail_text = normalize_spaces("".join(all_parts[-30:]))
            tail_text = (
                tail_text.replace("一次交清", "趸交")
                .replace("一次性付清", "趸交")
                .replace("一次性交付", "趸交")
            )
            found = [freq for freq in PAY_FREQ_ORDER if freq in tail_text]
            if found:
                return "，".join(found), sheet.title
    finally:
        workbook.close()
    return None


def extract_ci_pay_times(text: str) -> str | None:
    compact = normalize_spaces(text)
    if "重大疾病" not in compact and "重度疾病" not in compact:
        return None
    CI_BENEFIT = ["重大疾病保险金", "重度疾病保险金"]

    single_markers = ["本合同终止", "本项责任终止", "1次"]
    optional_multi_markers = ["选择多次给付", "附加多次给付", "可选择多次", "多次给付责任"]
    if any(marker in compact for marker in single_markers) and any(
        marker in compact for marker in optional_multi_markers
    ):
        numbers = [int(num) for num in re.findall(r"(\d+)次", compact) if int(num) > 1]
        if numbers:
            return f"1次（若选择多次给付责任，{numbers[0]}次）"

    _CN_NUM = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5,
               "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}
    m = re.search(r"(?:最多|至多|共|仅|只)(?:给付|赔付|赔偿).{0,10}?([一二三四五六七八九十]|\d+)次", compact)
    if m:
        raw = m.group(1)
        num = _CN_NUM.get(raw, raw)
        return f"{num}次"

    negative_markers = [
        "多次给付重大疾病保险金",
        "第二次重大疾病保险金",
        "第二次重疾",
        "多次重疾",
    ]
    has_negative = any(marker in compact for marker in negative_markers) or re.search(
        r"再次发生.{0,10}重大疾病", compact
    )
    positive = (any(k in compact for k in CI_BENEFIT) and "本合同终止" in compact) or (
        any(k in compact for k in ["重大疾病", "重度疾病"]) and "本项责任终止" in compact and "合同" in compact and "有效" in compact
    )
    if not has_negative and positive:
        return "1次"
    return None


def extract_ci_grouping(text: str) -> str | None:
    compact = normalize_spaces(text)
    if "分组" in compact:
        if "不分组" in compact:
            return "不分组"
        if "不同组" in compact or "分为" in compact:
            return "涉及分组"
    return None


def is_multi_ci_pay_times(value: str | None) -> bool | None:
    if not value:
        return None
    compact = normalize_spaces(value)
    if any(token in compact for token in ["多次", "多赔", "无限次", "无限"]):
        return True
    for num in re.findall(r"\d+", compact):
        if int(num) >= 2:
            return True
    if compact in {"1", "1次"}:
        return False
    if "次" in compact:
        return False
    return False


def extract_ci_count(text: str) -> str | None:
    compact = normalize_spaces(text)
    if "重大疾病" not in compact and "重度疾病" not in compact:
        return None
    m = re.search(r"重大疾病（共(\d+)种）", compact)
    if m:
        return f"{m.group(1)}种"
    m = re.search(r"重大疾病共有(\d+)种", compact)
    if m:
        return f"{m.group(1)}种"
    m = re.search(r"共计(\d+)种", compact)
    if m:
        return f"{m.group(1)}种"
    m = re.search(r"(\d+)种重大疾病", compact)
    if m:
        return f"{m.group(1)}种"
    m = re.search(r"共计([一二三四五六七八九十百零〇]+)种", compact)
    if m:
        val = chinese_to_int(m.group(1))
        if val is not None:
            return f"{val}种"
    return None


def should_skip_ci_count_block(block: dict) -> bool:
    title_path = "".join(block.get("title_path", []) or [])
    text = normalize_spaces(block.get("text", ""))
    compact = f"{title_path}{text}"
    skip_keywords = [
        "轻度疾病",
        "轻症",
        "中度疾病",
        "中症",
        "少儿特定",
    ]
    return any(keyword in compact for keyword in skip_keywords)


def extract_ci_count_from_disease_seq(blocks: list[dict]) -> str | None:
    max_seq = 0
    for block in blocks:
        if should_skip_ci_count_block(block):
            continue
        text = block.get("text", "")
        if not text:
            continue
        compact = normalize_spaces(text)
        if "重大疾病" not in compact and "重度疾病" not in compact and "恶性肿瘤" not in compact:
            # 序号段正文通常不再重复写“重大疾病”，但至少要避开明显非重疾病章节；
            # 这里仅用章节排除做约束，允许正文本身只保留病种列表。
            pass
        seqs = [int(num) for num in re.findall(r"(?<!\d)(\d{1,3})[\.、]\s*[^\n]{2,20}", text)]
        if seqs:
            max_seq = max(max_seq, max(seqs))
    # 兜底只接受明显像病种附表的大序号，避免把章节编号/责任条款编号误当成病种数。
    return f"{max_seq}种" if max_seq >= 50 else None


def extract_ci_count_from_section_numbering(blocks: list[dict]) -> str | None:
    section_no = None
    section_pattern = re.compile(r"^\s*(\d+)\s+重[大度]疾病(?:的)?定义")
    skip_keywords = ("轻度疾病", "中度疾病", "少儿特定")

    for idx, block in enumerate(blocks):
        text = (block.get("text") or "").strip()
        title_text = "".join(block.get("title_path", []) or []).strip()
        haystack = text or title_text
        if not haystack:
            continue
        match = section_pattern.search(haystack)
        if match:
            section_no = match.group(1)
            break

    if section_no is None:
        return None

    max_sub = 0
    sub_pattern = re.compile(rf"\b{re.escape(section_no)}\.(\d{{1,3}})\b")

    for block in blocks:
        text = (block.get("text") or "").strip()
        title_text = "".join(block.get("title_path", []) or []).strip()
        haystack = f"{title_text} {text}".strip()
        if not haystack:
            continue

        if any(keyword in haystack for keyword in skip_keywords):
            continue

        for raw in sub_pattern.findall(haystack):
            try:
                max_sub = max(max_sub, int(raw))
            except ValueError:
                continue

    return f"{max_sub}种" if max_sub >= 10 else None


def _find_fragmented_appendix_max_seq(blocks: list[dict], appendix_no: int, title_keyword: str) -> int:
    appendix_start = None
    appendix_end = len(blocks)
    appendix_tokens = {f"附录{appendix_no}", f"附录 {appendix_no}"}
    next_appendix_tokens = {f"附录{appendix_no + 1}", f"附录 {appendix_no + 1}"}

    for idx, block in enumerate(blocks):
        text = normalize_spaces(block.get("text", ""))
        if text.count("附录") > 1:
            continue
        if any(token in text for token in appendix_tokens) and title_keyword in text:
            appendix_start = idx
            break
        if text in appendix_tokens and idx + 1 < len(blocks):
            next_text = normalize_spaces(blocks[idx + 1].get("text", ""))
            if title_keyword in next_text:
                appendix_start = idx
                break

    if appendix_start is None:
        return 0

    for idx in range(appendix_start + 1, len(blocks)):
        text = normalize_spaces(blocks[idx].get("text", ""))
        if any(token in text for token in next_appendix_tokens):
            appendix_end = idx
            break

    max_seq = 0
    for idx in range(appendix_start, appendix_end - 1):
        cur_text = normalize_spaces(blocks[idx].get("text", ""))
        next_text = normalize_spaces(blocks[idx + 1].get("text", ""))
        if not re.fullmatch(r"\d{1,3}", cur_text):
            continue
        if not (2 <= len(next_text) <= 30):
            continue
        if re.search(r"[，。；：:（）()、\d]", next_text):
            continue
        max_seq = max(max_seq, int(cur_text))
    return max_seq


def extract_ci_count_from_fragmented_appendix_seq(blocks: list[dict]) -> str | None:
    main_count = _find_fragmented_appendix_max_seq(blocks, 3, "重大疾病定义")
    if main_count < 50:
        return None
    juvenile_count = _find_fragmented_appendix_max_seq(blocks, 5, "少儿特定疾病定义")
    if juvenile_count >= 5:
        return f"{main_count}种重大疾病；{juvenile_count}种少儿重大疾病"
    return f"{main_count}种"


def ci_count_confidence(text: str) -> float:
    compact = normalize_spaces(text)
    if any(pat in compact for pat in ["重大疾病（共", "重大疾病共有", "我们提供保障的重大疾病共有"]):
        return 0.95
    if "种重大疾病" in compact or "共计" in compact:
        return 0.86
    return 0.8


def extract_candidates(blocks: list[dict], product_id: str | None = None) -> dict:
    results: dict = {}
    grouping_match: dict | None = None
    text_blocks: list[dict] = []
    for idx, block in enumerate(blocks):
        if block["block_type"] not in {"paragraph", "title", "list"}:
            continue
        text = block["text"]
        pay_period_text = text
        title_text = " ".join(block.get("title_path", []))
        if any(keyword in text or keyword in title_text for keyword in PAY_PERIOD_TITLES):
            next_block = blocks[idx + 1] if idx + 1 < len(blocks) else None
            if next_block and next_block.get("block_type") in {"paragraph", "title", "list"}:
                pay_period_text = f"{text} {next_block.get('text', '')}".strip()
        text_blocks.append(block)

        add_candidate(results, "投保年龄", extract_age(text), block, 0.95 if "投保年龄" in text else 0.82, "rule: age_pattern")
        _ip_val = extract_insurance_period(text)
        _ip_conf = 0.97 if any("保险期间" in t for t in block.get("title_path", [])) else 0.93
        add_candidate(results, "保险期间", _ip_val, block, _ip_conf, "rule: insurance_period_pattern")
        add_candidate(results, "交费期间", extract_pay_period(pay_period_text), block, 0.9, "rule: pay_period_pattern")
        add_candidate(results, "交费频率", extract_pay_frequency(text), block, 0.9, "rule: pay_frequency_pattern")

        waiting_raw, waiting_simple = extract_waiting(text)
        waiting_conf = waiting_block_confidence(block)
        add_candidate(results, "等待期", waiting_raw, block, waiting_conf, "rule: waiting_period_pattern")
        add_candidate(results, "等待期（简化）", waiting_simple, block, waiting_conf, "rule: waiting_period_simple")

        grace_val = extract_duration_days(text, "宽限期")
        add_candidate(results, "宽限期", grace_val, block, 0.94, "rule: grace_period_pattern")
        if grace_val and results.get("宽限期", {}).get("block_id") == block["block_id"]:
            results["宽限期"]["evidence_text"] = narrow_evidence(text, "宽限期")

        cooling_val = extract_duration_days(text, "犹豫期", max_days=30)
        add_candidate(results, "犹豫期", cooling_val, block, 0.94, "rule: cooling_off_pattern")
        if cooling_val and results.get("犹豫期", {}).get("block_id") == block["block_id"]:
            results["犹豫期"]["evidence_text"] = narrow_evidence(text, "犹豫期")
        add_candidate(results, "重疾赔付次数", extract_ci_pay_times(text), block, 0.82, "rule: ci_pay_times")
        grouping_value = extract_ci_grouping(text)
        if grouping_value:
            candidate = {
                "coverage_name": "重疾分组",
                "value": grouping_value,
                "confidence": 0.75,
                "note": "rule: ci_grouping",
                "block_id": block["block_id"],
                "page": block["page"],
                "evidence_text": block["text"],
            }
            if grouping_match is None or candidate["confidence"] > grouping_match["confidence"]:
                grouping_match = candidate
        if not should_skip_ci_count_block(block):
            count_value = extract_ci_count(text)
            add_candidate(results, "重疾数量", count_value, block, ci_count_confidence(text), "rule: ci_count")

    # 跨block兜底：重疾赔付次数单次判定
    if "重疾赔付次数" not in results:
        multi_markers = ["多次给付重大疾病", "第二次重大疾病", "多次重疾", "再次发生.*重大疾病"]
        all_text = " ".join(b["text"] for b in blocks)
        has_termination = any(k in all_text for k in ["本合同终止", "合同效力终止", "主合同终止", "本主险合同效力终止"])
        has_ci_termination = any(
            any(k in normalize_spaces(b["text"]) for k in ["本合同终止", "合同效力终止", "主合同终止", "本主险合同效力终止"])
            and any(ci in normalize_spaces(b["text"]) for ci in ["重大疾病保险金", "重度疾病保险金"])
            for b in blocks
        )
        has_ci = any(k in all_text for k in ["重大疾病保险金", "重度疾病保险金"])
        has_multi = any(m in all_text for m in multi_markers) or re.search(r"再次发生.{0,10}重大疾病", all_text)
        has_child_ci = any("少儿重大疾病" in b["text"] or "少儿特定疾病" in b["text"] for b in blocks)
        if has_ci and has_ci_termination and not has_multi and not has_child_ci:
            evidence_block = next(
                (b for b in blocks if any(k in normalize_spaces(b["text"]) for k in ["本合同终止", "合同效力终止", "主合同终止", "本主险合同效力终止"])
                 and any(ci in normalize_spaces(b["text"]) for ci in ["重大疾病保险金", "重度疾病保险金"])),
                next((b for b in blocks if any(k in normalize_spaces(b["text"]) for k in ["本合同终止", "合同效力终止", "主合同终止", "本主险合同效力终止"])), blocks[0])
            )
            add_candidate(results, "重疾赔付次数", "1次", evidence_block, 0.78, "rule: cross_block_single_pay")

    # 跨block兜底：可选多次赔付（如851A"可选部分的第二次重大疾病保险金"）
    if "重疾赔付次数" not in results:
        has_optional_part = any("可选部分" in b["text"] for b in blocks)
        has_second_pay = any(
            any(k in b["text"] for k in ["第二次重大疾病", "第二次重度疾病"])
            for b in blocks
        )
        if has_optional_part and has_second_pay:
            evidence_block = next(
                (b for b in blocks if any(k in b["text"] for k in ["第二次重大疾病", "第二次重度疾病"])),
                blocks[0]
            )
            add_candidate(results, "重疾赔付次数", "1次（若选择多次给付责任，2次）",
                          evidence_block, 0.78, "rule: cross_block_optional_multi")

    # 保险期间兜底：从合同名称推断终身
    if "保险期间" not in results:
        for b in blocks:
            if b.get("page", 99) <= 2 and re.search(r"终身.{0,8}(?:重大疾病|疾病)保险", b.get("text", "")):
                add_candidate(results, "保险期间", "终身", b, 0.80, "rule: inferred_from_contract_name")
                break

    pay_times_value = results.get("重疾赔付次数", {}).get("value")
    multi_pay = is_multi_ci_pay_times(pay_times_value)
    if multi_pay is None:
        # Cannot determine grouping without knowing pay-times; leave as missing
        pass
    elif multi_pay is False:
        pay_times_block = results.get("重疾赔付次数")
        set_candidate(
            results,
            "重疾分组",
            "不涉及",
            pay_times_block,
            0.9,
            "rule: ci_grouping_from_pay_times_single",
        )
    elif pay_times_value and "若选择多次给付责任" in pay_times_value:
        pay_times_block = results.get("重疾赔付次数")
        set_candidate(
            results,
            "重疾分组",
            "不分组",
            pay_times_block,
            0.75,
            "rule: ci_grouping_optional_multi",
        )
    elif grouping_match is not None:
        results["重疾分组"] = grouping_match
    # else: multi-pay but no grouping info found → leave as missing

    # 等待期跨block升级：抽到纯天数但全文存在"意外无等待期"声明时，升级为完整格式
    if "等待期" in results:
        waiting_val = results["等待期"].get("value", "")
        if waiting_val and "意外" not in waiting_val:
            all_compact = normalize_spaces(" ".join(b["text"] for b in blocks))
            has_no_waiting_accident = (
                any("意外" in b["text"] and "无等待期" in b["text"] for b in blocks)
                or any("意外" in b["text"] and "不设等待期" in b["text"] for b in blocks)
                or re.search(r"意外(?:伤害)?[^。]{0,20}(?:无等待期|不设等待期|等待期为0[天日])", all_compact)
                or re.search(r"(?:无等待期|不设等待期)[^。]{0,20}意外(?:伤害)?", all_compact)
                or re.search(r"意外(?:伤害)?[^。]{0,30}不受[^。]{0,15}等待期", all_compact)
                or re.search(r"因意外(?:伤害)?[^。]{0,20}0[天日]", all_compact)
            )
            if has_no_waiting_accident:
                new_waiting_val = f"非意外{waiting_val}，意外0天"
                results["等待期"]["value"] = new_waiting_val
                # 找到触发升级的证据 block，同步 block_id/page/evidence_text/note
                upgrade_block = None
                for b in blocks:
                    bt = normalize_spaces(b["text"])
                    if (("意外" in bt and "无等待期" in bt)
                            or ("意外" in bt and "不设等待期" in bt)
                            or re.search(r"意外(?:伤害)?[^。]{0,20}(?:无等待期|不设等待期|等待期为0[天日])", bt)
                            or re.search(r"(?:无等待期|不设等待期)[^。]{0,20}意外(?:伤害)?", bt)
                            or re.search(r"意外(?:伤害)?[^。]{0,30}不受[^。]{0,15}等待期", bt)
                            or re.search(r"因意外(?:伤害)?[^。]{0,20}0[天日]", bt)):
                        upgrade_block = b
                        break
                if upgrade_block is not None:
                    results["等待期"]["block_id"] = upgrade_block.get("block_id", results["等待期"].get("block_id"))
                    results["等待期"]["page"] = upgrade_block.get("page", results["等待期"].get("page"))
                    results["等待期"]["evidence_text"] = upgrade_block["text"][:200]
                    results["等待期"]["note"] = "cross-block upgrade: accident_no_waiting"
                # 同步升级 等待期（简化）的 block 引用（值保持简化形式 N天）
                if "等待期（简化）" in results and upgrade_block is not None:
                    results["等待期（简化）"]["block_id"] = upgrade_block.get("block_id", results["等待期（简化）"].get("block_id"))
                    results["等待期（简化）"]["page"] = upgrade_block.get("page", results["等待期（简化）"].get("page"))
                    results["等待期（简化）"]["evidence_text"] = upgrade_block["text"][:200]
                    results["等待期（简化）"]["note"] = "cross-block upgrade: accident_no_waiting"

    # 等待期跨block兜底：旧式条款用"X日内...若因意外...不受限制"表达等待期，无"等待期"关键词
    if "等待期" not in results:
        for b in blocks:
            compact_b = normalize_spaces(b["text"])
            m = re.search(r"之日起([一二三四五六七八九十百零〇]+|\d+)[日天]内", compact_b)
            if m and re.search(r"若因意外伤害.*不受", compact_b):
                raw = m.group(1)
                val = chinese_to_int(raw) if not raw.isdigit() else int(raw)
                if val:
                    simplified = f"{val}天"
                    add_candidate(results, "等待期", f"非意外{simplified}，意外0天", b, 0.75, "rule: waiting_embedded_coverage_clause")
                    add_candidate(results, "等待期（简化）", simplified, b, 0.75, "rule: waiting_embedded_coverage_clause")
                break

    if "重疾数量" not in results:
        section_count = extract_ci_count_from_section_numbering(blocks)
        if section_count:
            add_candidate(results, "重疾数量", section_count, blocks[0] if blocks else None, 0.88, "rule: ci_count_section_numbering")

    if "重疾数量" not in results:
        fragmented_count = extract_ci_count_from_fragmented_appendix_seq(blocks)
        if fragmented_count:
            add_candidate(results, "重疾数量", fragmented_count, blocks[0] if blocks else None, 0.86, "rule: ci_count_fragmented_appendix_seq")

    if "重疾数量" not in results:
        seq_count = extract_ci_count_from_disease_seq(blocks)
        if seq_count:
            add_candidate(results, "重疾数量", seq_count, blocks[0] if blocks else None, 0.84, "rule: ci_count_from_disease_seq")

    return results


def pay_period_specificity(value: str) -> int:
    compact = normalize_spaces(value)
    score = 0
    if "趸交" in compact:
        score += 1
    for raw in re.findall(r"(\d+(?:/\d+)*)年交", compact):
        score += len([part for part in raw.split("/") if part])
    if "交至" in compact:
        score += 10
    return score


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract Tier A rule-based candidates from blocks.")
    parser.add_argument("manifest_pos", nargs="?")
    parser.add_argument("output_pos", nargs="?")
    parser.add_argument("--manifest")
    parser.add_argument("--blocks-dir")
    parser.add_argument("--output")
    args = parser.parse_args()

    manifest_path = args.manifest or args.manifest_pos
    output_path_arg = args.output or args.output_pos
    blocks_dir_arg = args.blocks_dir or str(Path(__file__).resolve().parents[1] / "data" / "blocks")

    if not manifest_path or not output_path_arg:
        parser.error("must provide manifest and output, either as positional args or with --manifest/--output")

    manifest = load_manifest_compatible(Path(manifest_path))
    rows = []
    for item in manifest:
        if item.get("phase1_eligible") is False:
            continue
        if item.get("status") == "phase1_excluded":
            continue
        blocks_path = Path(blocks_dir_arg) / f"{item['product_id']}_blocks.json"
        if not blocks_path.exists():
            degraded_path = Path(blocks_dir_arg) / f"{item['product_id']}_blocks_degraded.json"
            if not degraded_path.exists():
                continue
            blocks_path = degraded_path
        _, blocks = load_blocks_compatible(blocks_path)
        blocks = tag_blocks(blocks)
        extracted = extract_candidates(blocks, item["product_id"])
        shuomingshu_blocks_path = Path(blocks_dir_arg) / f"{item['product_id']}_说明书_blocks.json"
        if shuomingshu_blocks_path.exists() and ("投保年龄" not in extracted or "交费期间" not in extracted):
            _, shuomingshu_blocks = load_blocks_compatible(shuomingshu_blocks_path)
            shuomingshu_blocks = tag_blocks(shuomingshu_blocks)
            extra = extract_candidates(shuomingshu_blocks, item["product_id"])
            if "投保年龄" not in extracted and "投保年龄" in extra:
                age_candidate = extra["投保年龄"].copy()
                age_candidate["note"] = "rule: age_from_shuomingshu"
                age_candidate["source_type"] = "product_brochure"
                extracted["投保年龄"] = age_candidate
            if "交费期间" not in extracted and "交费期间" in extra:
                pay_period_candidate = extra["交费期间"].copy()
                pay_period_candidate["note"] = "rule: pay_period_shuomingshu"
                pay_period_candidate["source_type"] = "product_brochure"
                extracted["交费期间"] = pay_period_candidate
        if "投保年龄" not in extracted:
            age_from_rate = extract_age_from_rate_source(item)
            if age_from_rate:
                extracted["投保年龄"] = {
                    "coverage_name": "投保年龄",
                    "value": age_from_rate,
                    "confidence": 0.8,
                    "note": "rule: age_from_rate_table_header",
                    "block_id": None,
                    "page": None,
                    "evidence_text": None,
                }
        if "交费期间" not in extracted or "交费频率" not in extracted:
            rate_pdf_info = extract_pay_info_from_rate_pdf(item)
            if "交费期间" not in extracted and rate_pdf_info.get("交费期间"):
                extracted["交费期间"] = {
                    "coverage_name": "交费期间",
                    "value": rate_pdf_info["交费期间"],
                    "confidence": 0.8,
                    "note": "rule: rate_pdf_pay_period",
                    "block_id": None,
                    "page": None,
                    "evidence_text": None,
                }
            if "交费频率" not in extracted and rate_pdf_info.get("交费频率"):
                extracted["交费频率"] = {
                    "coverage_name": "交费频率",
                    "value": rate_pdf_info["交费频率"],
                    "confidence": 0.8,
                    "note": "rule: rate_pdf_pay_frequency",
                    "block_id": None,
                    "page": None,
                    "evidence_text": None,
                }
        # 若已有交费期间候选但仅有"趸交"（无年期），且来源为说明书，则也尝试费率表补全
        _pp_existing = extracted.get("交费期间")
        _pp_only_lump = (
            _pp_existing is not None
            and _pp_existing.get("value") == "趸交"
            and "shuomingshu" in _pp_existing.get("note", "")
        )
        _pp_from_shuomingshu = (
            _pp_existing is not None
            and "shuomingshu" in (_pp_existing.get("note") or "")
        )
        if "交费期间" not in extracted or _pp_only_lump or _pp_from_shuomingshu:
            rate_xlsx = locate_rate_xlsx(item)
            if rate_xlsx and rate_xlsx.exists():
                try:
                    from openpyxl import load_workbook as _lw

                    _LUMP_SYNONYMS = {"趸交", "一次性付清", "一次交清", "一次性交付"}

                    def _fmt_periods(periods: list) -> str | None:
                        has_lump = any(p in _LUMP_SYNONYMS for p in periods)
                        yrs = sorted({int(m2.group(1)) for p in periods for m2 in [re.fullmatch(r"(\d+)年交", p)] if m2})
                        parts = (["趸交"] if has_lump else []) + (["/".join(str(y) for y in yrs) + "年交"] if yrs else [])
                        return "，".join(parts) if parts else None

                    wb = _lw(rate_xlsx, read_only=True, data_only=True)
                    sheet = next(
                        (s for s in wb.sheetnames if not any(k in s for k in ("次标准", "次标准体", "优选体"))),
                        wb.sheetnames[0],
                    )
                    ws = wb[sheet]
                    pay_periods_found: list[str] = []
                    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
                        for cell in row:
                            if cell is None:
                                continue
                            val = str(cell).strip()
                            if val in _LUMP_SYNONYMS:
                                pay_periods_found.append("趸交")
                            m = re.fullmatch(r"(\d+)年交", val)
                            if m:
                                pay_periods_found.append(val)
                    if pay_periods_found:
                        formatted = _fmt_periods(pay_periods_found)
                        if formatted:
                            should_replace = (
                                "交费期间" not in extracted
                                or _pp_only_lump
                                or (
                                    _pp_from_shuomingshu
                                    and pay_period_specificity(formatted)
                                    > pay_period_specificity(_pp_existing.get("value") or "")
                                )
                            )
                            if should_replace:
                                extracted["交费期间"] = {
                                    "coverage_name": "交费期间",
                                    "value": formatted,
                                    "confidence": 0.82,
                                    "note": f"rule: rate_xlsx_header[{sheet}]",
                                    "block_id": None,
                                    "page": None,
                                    "evidence_text": str(rate_xlsx),
                                }
                except Exception:
                    pass
        if "交费频率" not in extracted:
            rate_freq = extract_pay_frequency_from_rate_xlsx(item)
            if rate_freq:
                value, sheet_name = rate_freq
                extracted["交费频率"] = {
                    "coverage_name": "交费频率",
                    "value": value,
                    "confidence": 0.9,
                    "note": f"rule: pay_freq_rate_table[{sheet_name}]",
                    "block_id": None,
                    "page": None,
                    "evidence_text": None,
                }
        # 趸交联动：交费期间含趸交但交费频率未含趸交时，补充趸交
        if "交费频率" in extracted and "交费期间" in extracted:
            pp_val = extracted["交费期间"].get("value") or ""
            freq_cand = extracted["交费频率"]
            freq_val = freq_cand.get("value") or ""
            if "趸交" in pp_val and "趸交" not in freq_val:
                freq_cand = dict(freq_cand)
                freq_cand["value"] = "趸交，" + freq_val
                freq_cand["note"] = (freq_cand.get("note") or "") + "+趸交(linked from 交费期间)"
                extracted["交费频率"] = freq_cand
        # 无频率兜底：完全没有交费频率时，按交费期间推断最小频率
        if "交费频率" not in extracted and "交费期间" in extracted:
            _pp = extracted["交费期间"].get("value") or ""
            _has_lump = "趸交" in _pp
            _has_year = any(k in _pp for k in ["年交", "交至"])
            if _has_lump or _has_year:
                _inferred_parts = []
                if _has_lump:
                    _inferred_parts.append("趸交")
                if _has_year:
                    _inferred_parts.append("年交")
                extracted["交费频率"] = {
                    "coverage_name": "交费频率",
                    "value": "，".join(_inferred_parts),
                    "confidence": 0.6,
                    "note": "rule: pay_freq_inferred_from_pay_period",
                    "block_id": None,
                    "page": None,
                    "evidence_text": None,
                }
        structured_table_path = locate_structured_table_json(item, "raw_rate")
        if structured_table_path and structured_table_path.exists():
            try:
                structured_candidates = build_candidates_from_structured_table(structured_table_path)
            except Exception:
                structured_candidates = {}
            for field in ("保险期间", "交费期间", "交费频率"):
                if field not in structured_candidates:
                    continue
                if field not in extracted:
                    extracted[field] = structured_candidates[field]
                    continue
                if field == "交费期间":
                    existing = extracted[field]
                    sc = structured_candidates[field]
                    existing_spec = pay_period_specificity(existing.get("value") or "")
                    sc_spec = pay_period_specificity(sc.get("value") or "")
                    if sc_spec > existing_spec:
                        is_xlsx = sc.get("parse_quality") in ("xlsx", "xls")
                        from_shuomingshu = "shuomingshu" in (existing.get("note") or "")
                        # Excel sources always trusted; PDF sources only when
                        # existing extraction looks incomplete (specificity ≤ 2)
                        if is_xlsx or from_shuomingshu or existing_spec <= 2:
                            extracted[field] = sc
        rows.append(
            {
                "product_id": item["product_id"],
                "db_product_id": item.get("db_product_id"),
                "product_name": item["product_name"],
                "source_blocks": str(blocks_path),
                "candidates": [extracted[field] for field in TARGET_FIELDS if field in extracted],
                "missing_fields": [field for field in TARGET_FIELDS if field not in extracted],
            }
        )

    output_path = Path(output_path_arg)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {len(rows)} products to {output_path}")


if __name__ == "__main__":
    main()
