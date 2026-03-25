#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PACKAGE_ROOT = Path(__file__).resolve().parent
PARENT_ROOT = PACKAGE_ROOT.parent
if str(PARENT_ROOT) not in sys.path:
    sys.path.append(str(PARENT_ROOT))

from rate_standardization.rate_field_dicts import (
    CHINESE_TO_INTERNAL_FIELD,
    COMPANY_ID_BY_NAME,
    COMPANY_NAME_ALIASES,
    CORE_CMB_PRODUCT_RATE_FIELDS,
    DATABASE_DEFAULTS,
    FINAL_RATE_HEADERS_3,
    GENDER_NORMALIZATION_MAP,
    INSURANCE_PERIOD_NORMALIZATION_MAP,
    INTERNAL_TO_DB_FIELD,
    PAYMENT_YEARS_NORMALIZATION_MAP,
    PAY_TYPE_NORMALIZATION_MAP,
)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_company_name(value: Any) -> str:
    text = normalize_text(value)
    return COMPANY_NAME_ALIASES.get(text, text)


def normalize_age(value: Any) -> int:
    text = normalize_text(value)
    return int(text) if text.isdigit() else 0


def normalize_payment_years(value: Any) -> str:
    text = normalize_text(value)
    return PAYMENT_YEARS_NORMALIZATION_MAP.get(text, text)


def normalize_insurance_period(value: Any) -> str:
    text = normalize_text(value)
    return INSURANCE_PERIOD_NORMALIZATION_MAP.get(text, text)


def normalize_gender(value: Any) -> int:
    text = normalize_text(value)
    return GENDER_NORMALIZATION_MAP.get(text, 0)


def normalize_pay_type(value: Any) -> int:
    text = normalize_text(value)
    return PAY_TYPE_NORMALIZATION_MAP.get(text, 0)


def normalize_amount(value: Any) -> int:
    text = normalize_text(value).replace(",", "")
    return int(text) if text.isdigit() else 1000


def normalize_rate(value: Any) -> float:
    text = normalize_text(value).replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def first_sheet_rows(path: Path) -> tuple[list[str], list[tuple[Any, ...]]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [normalize_text(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    return headers, rows


def build_header_index(headers: list[str]) -> dict[str, int]:
    return {header: idx for idx, header in enumerate(headers) if header}


def cell(row: tuple[Any, ...], header_index: dict[str, int], header: str) -> Any:
    idx = header_index.get(header)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def to_db_row(row: tuple[Any, ...], header_index: dict[str, int], source_file: str) -> dict[str, Any]:
    company_name = normalize_company_name(cell(row, header_index, "公司"))
    company_id = normalize_text(cell(row, header_index, "公司编号")) or COMPANY_ID_BY_NAME.get(company_name, "")
    product_id = normalize_text(cell(row, header_index, "产品编号"))
    product_name = normalize_text(cell(row, header_index, "产品"))

    db_row: dict[str, Any] = {
        "product_id": product_id,
        "amount": normalize_amount(cell(row, header_index, "固定保额/份数")),
        "rate": normalize_rate(cell(row, header_index, "费率")),
        "age": normalize_age(cell(row, header_index, "投保年龄")),
        "gender": normalize_gender(cell(row, header_index, "性别")),
        "payment_years": normalize_payment_years(cell(row, header_index, "缴费期间")),
        "insurance_period": normalize_insurance_period(cell(row, header_index, "保险期间")),
        "pay_type": normalize_pay_type(cell(row, header_index, "交费方式")),
        "company_id": company_id,
        "company_name": company_name,
        "product_name": product_name,
        "rate_unit": "per_1000_amount",
        "source_type": "processed_rate",
        "source_file": source_file,
    }

    for key, value in DATABASE_DEFAULTS.items():
        db_row.setdefault(key, value)
    return db_row


def parse_processed_rate(path: Path) -> dict[str, Any]:
    headers, rows = first_sheet_rows(path)
    header_index = build_header_index(headers)
    missing_headers = [header for header in FINAL_RATE_HEADERS_3 if header not in header_index and header not in {"年金领取年龄"}]

    db_rows = [to_db_row(row, header_index, str(path)) for row in rows if normalize_text(cell(row, header_index, "产品编号"))]
    preview = db_rows[:5]

    return {
        "source_file": str(path),
        "header_count": len(headers),
        "headers": headers,
        "missing_headers": missing_headers,
        "row_count": len(db_rows),
        "core_db_fields": CORE_CMB_PRODUCT_RATE_FIELDS,
        "rows": db_rows,
        "preview_rows": preview,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse processed rate xlsx into cmb_product_rate-aligned rows.")
    parser.add_argument("--input", required=True, help="Processed rate xlsx path")
    parser.add_argument("--output", required=True, help="Output json path")
    args = parser.parse_args()

    result = parse_processed_rate(Path(args.input))
    Path(args.output).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"row_count": result["row_count"], "output": args.output}, ensure_ascii=False))


if __name__ == "__main__":
    main()
