#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PACKAGE_ROOT = Path(__file__).resolve().parent
PARENT_ROOT = PACKAGE_ROOT.parent
if str(PARENT_ROOT) not in sys.path:
    sys.path.append(str(PARENT_ROOT))

from rate_standardization.rate_field_dicts import (
    DATABASE_DEFAULTS,
    EXCLUDED_SHEET_KEYWORDS,
    GENDER_NORMALIZATION_MAP,
    PAYMENT_YEARS_NORMALIZATION_MAP,
    encode_insurance_period,
    encode_payment_years,
)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_rate(value: Any) -> float | None:
    text = normalize_text(value).replace(",", "")
    if text in {"", "-"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_age(value: Any) -> int | None:
    text = normalize_text(value)
    if not text.isdigit():
        return None
    return int(text)


def normalize_pay_period(text: str) -> str:
    if text == "一次性付清":
        return "趸交"
    return PAYMENT_YEARS_NORMALIZATION_MAP.get(text, text)


def select_sheet(workbook, sheet_name: str | None) -> str:
    all_sheets = workbook.sheetnames
    if sheet_name is not None:
        if sheet_name not in all_sheets:
            raise ValueError(f"指定的 sheet '{sheet_name}' 不存在，可用: {all_sheets}")
        return sheet_name
    if "费率表" in all_sheets:
        return "费率表"
    candidates = [s for s in all_sheets if not any(kw in s for kw in EXCLUDED_SHEET_KEYWORDS)]
    if len(candidates) == 0:
        raise ValueError(f"所有 sheet 均被排除，请用 --sheet 显式指定。sheets: {all_sheets}")
    if len(candidates) > 1:
        print(f"[警告] 多个候选 sheet，自动选择第一个: {candidates[0]}，其余: {candidates[1:]}")
    return candidates[0]


def get_merged_value_map(ws, row_num: int) -> dict[int, str]:
    raw = {cell.column: cell.value for cell in ws[row_num] if cell.value is not None}
    result: dict[int, str] = {}
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row_num <= merged_range.max_row:
            top_left_val = ws.cell(merged_range.min_row, merged_range.min_col).value
            if top_left_val is not None:
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    result[col] = normalize_text(top_left_val)
    for col, val in raw.items():
        if col not in result:
            result[col] = normalize_text(val)
    return result


def detect_template(ws) -> str:
    row5_vals = {normalize_text(v.value) for v in ws[5] if v.value is not None}
    if "男性" in row5_vals or "女性" in row5_vals:
        return "A"
    return "C"


def infer_pay_type(pay_period: str) -> int:
    return 0 if pay_period == "趸交" else 12


def build_row(
    *,
    product_id: str,
    company_id: str,
    company_name: str,
    product_name: str,
    source_file: str,
    age: int,
    gender_text: str,
    pay_period_text: str,
    rate: float,
) -> dict[str, Any]:
    pay_period = normalize_pay_period(pay_period_text)
    row: dict[str, Any] = {
        "product_id": product_id,
        "age": age,
        "gender": GENDER_NORMALIZATION_MAP.get(gender_text, 0),
        **encode_payment_years(pay_period),
        **encode_insurance_period("终身"),
        "rate": rate,
        "pay_type": infer_pay_type(pay_period),
        "amount": 1000,
        "company_id": company_id,
        "company_name": company_name,
        "product_name": product_name,
        "rate_unit": "per_1000_amount",
        "source_type": "raw_rate",
        "source_file": source_file,
    }
    for key, value in DATABASE_DEFAULTS.items():
        row.setdefault(key, value)
    return row


def parse_template_a(ws, meta: dict[str, str]) -> list[dict[str, Any]]:
    outer = get_merged_value_map(ws, 5)
    inner = get_merged_value_map(ws, 6)
    rows: list[dict[str, Any]] = []
    for excel_row in range(7, ws.max_row + 1):
        age = normalize_age(ws.cell(excel_row, 2).value)
        if age is None:
            continue
        for col in range(3, ws.max_column + 1):
            gender_text = normalize_text(outer.get(col))
            pay_period_text = normalize_text(inner.get(col))
            if gender_text not in GENDER_NORMALIZATION_MAP or not pay_period_text:
                continue
            rate = normalize_rate(ws.cell(excel_row, col).value)
            if rate is None:
                continue
            rows.append(
                build_row(
                    product_id=meta["product_id"],
                    company_id=meta["company_id"],
                    company_name=meta["company_name"],
                    product_name=meta["product_name"],
                    source_file=meta["source_file"],
                    age=age,
                    gender_text=gender_text,
                    pay_period_text=pay_period_text,
                    rate=rate,
                )
            )
    return rows


def parse_template_c(ws, meta: dict[str, str]) -> list[dict[str, Any]]:
    outer = get_merged_value_map(ws, 6)
    inner = get_merged_value_map(ws, 7)
    rows: list[dict[str, Any]] = []
    for excel_row in range(8, ws.max_row + 1):
        age = normalize_age(ws.cell(excel_row, 2).value)
        if age is None:
            continue
        for col in range(3, ws.max_column + 1):
            pay_period_text = normalize_text(outer.get(col))
            gender_text = normalize_text(inner.get(col))
            if gender_text not in GENDER_NORMALIZATION_MAP or not pay_period_text:
                continue
            rate = normalize_rate(ws.cell(excel_row, col).value)
            if rate is None:
                continue
            rows.append(
                build_row(
                    product_id=meta["product_id"],
                    company_id=meta["company_id"],
                    company_name=meta["company_name"],
                    product_name=meta["product_name"],
                    source_file=meta["source_file"],
                    age=age,
                    gender_text=gender_text,
                    pay_period_text=pay_period_text,
                    rate=rate,
                )
            )
    return rows


def parse_raw_rate(
    path: Path,
    product_id: str,
    company_id: str,
    company_name: str,
    product_name: str,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    selected_sheet = select_sheet(workbook, sheet_name)
    ws = workbook[selected_sheet]
    template_type = detect_template(ws)
    meta = {
        "product_id": product_id,
        "company_id": company_id,
        "company_name": company_name,
        "product_name": product_name,
        "source_file": str(path),
    }
    if template_type == "A":
        rows = parse_template_a(ws, meta)
    else:
        rows = parse_template_c(ws, meta)
    return {
        "source_file": str(path),
        "sheet_name": selected_sheet,
        "row_count": len(rows),
        "rows": rows,
        "preview_rows": rows[:5],
        "template_type": template_type,
        "missing_note": None,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse raw rate xlsx into DB-ready rows.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--product-id", required=True)
    parser.add_argument("--company-id", required=True)
    parser.add_argument("--company-name", required=True)
    parser.add_argument("--product-name", required=True)
    parser.add_argument("--sheet")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    result = parse_raw_rate(
        path=Path(args.input).expanduser(),
        product_id=args.product_id,
        company_id=args.company_id,
        company_name=args.company_name,
        product_name=args.product_name,
        sheet_name=args.sheet,
    )
    output_path = Path(args.output).expanduser()
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"row_count": result["row_count"], "sheet_name": result["sheet_name"], "output": str(output_path)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
